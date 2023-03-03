import os
import string
import openpyxl
from openpyxl.styles import Font, colors
from random import sample
import time
import numpy
import json
import glob
import natsort

# Font color setting
FAIL = "00FF0000"

# file read setting
TRAIN = 75
TEST = 25
DATA_PATH = 'Datafile_0222/'
SHEET_NAME = 'Data_1'

START_ROW = 24
TESTNUM_COLUMN = 2
TESTNAME_COLUMN = 3

# fail rate setting
DATA_FILE_NAME = []
DATA_FILE_LIST = []
testFail = {}
trainFail = {}
Similarity = [{} for i in range(TEST)]


def fail_calculate_TRAIN(DATA_FILE_LIST):
    # read testfile
    for file_path in DATA_FILE_LIST:
        DATA_FILE_NAME.append(os.path.basename(file_path))
    print("\n##############  1. Estimate Training File Fail Rate Process ##############\n")
    # In Train case calculate which TestName is fail
    for a in range(TRAIN):
        wb = openpyxl.load_workbook(DATA_FILE_LIST[a])
        print(" Estimating training file " +
              DATA_FILE_NAME[a])
        s2 = wb[SHEET_NAME]
        # save the list of testing testname
        targetList = []
        for i in range(START_ROW, s2.max_row+1):
            cell_obj = s2.cell(row=i, column=TESTNUM_COLUMN)
            if cell_obj.font.color is not None and cell_obj.font.color.rgb == FAIL:
                targetList.append(s2.cell(row=i, column=TESTNAME_COLUMN).value)
        trainFail[DATA_FILE_NAME[a]] = targetList
    return trainFail


def fail_calculate_TEST(DATA_FILE_LIST):
    # read testfile
    for file_path in DATA_FILE_LIST:
        DATA_FILE_NAME.append(os.path.basename(file_path))
    print("\n##############  2. Estimate testing file fail rate process ##############\n")
    # In Test case calculate which TestName is fail
    for b in range(TEST):
        wb = openpyxl.load_workbook(DATA_FILE_LIST[b+TRAIN])
        s2 = wb[SHEET_NAME]
        print(" Estimating training file " +
              DATA_FILE_NAME[b+TRAIN])
        # 把target的testname存進去list
        targetList = []
        for i in range(START_ROW, s2.max_row+1):
            cell_obj = s2.cell(row=i, column=TESTNUM_COLUMN)
            if cell_obj.font.color is not None and cell_obj.font.color.rgb == FAIL:
                targetList.append(s2.cell(row=i, column=TESTNAME_COLUMN).value)
        testFail[DATA_FILE_NAME[b+TRAIN]] = targetList
    return testFail


def similar(trainFail, testFail, DATA_FILE_NAME):
    Rate = {}
    similar = []

    print("\n##############  3. Estimate the similarity between the Training/Testing file #####################")

    for i in range(TEST):
        rateDict = {}
        for j in range(TRAIN):
            cal_rate = 0
            count = 0
            for k in range(len(testFail.get(DATA_FILE_NAME[i+TRAIN]))):
                if testFail.get(DATA_FILE_NAME[i+TRAIN])[k] in (trainFail.get(DATA_FILE_NAME[j])):
                    count += 1
                else:
                    pass
            cal_rate = count / len(numpy.unique((testFail.get(DATA_FILE_NAME[i+TRAIN]) +
                                                (trainFail.get(DATA_FILE_NAME[j])))))
            rateDict[DATA_FILE_NAME[j]] = cal_rate
        similar = sorted(rateDict.items(),
                         key=lambda item: item[1], reverse=True)
        Rate[DATA_FILE_NAME[i]] = similar
    print("\n  Estimation of similarity finished  \n")
    jsonFile = open("Result/Baseline_similars_result.json", "w")
    jsonFile.write(json.dumps(Rate, indent=2))
    jsonFile.close()


def main():
    # time calculate
    start_time = time.time()

    path = 'Result'
    if not os.path.isdir(path):
        os.makedirs(path)

    #================== Sort the file names first ==================#
    DATA_FILE_LIST = glob.glob(os.path.join(DATA_PATH, "*.xlsx"))
    DATA_FILE_LIST = natsort.natsorted(DATA_FILE_LIST)
    Training_Result = fail_calculate_TRAIN(DATA_FILE_LIST)
    Test_Result = fail_calculate_TEST(DATA_FILE_LIST)
    similar(Training_Result, Test_Result, DATA_FILE_NAME)
    print("*********************************************\n*                                           *\n*         %s mins            *\n*                                           *\n*********************************************\n" % ((time.time() - start_time) / 60))


if __name__ == "__main__":
    main()
